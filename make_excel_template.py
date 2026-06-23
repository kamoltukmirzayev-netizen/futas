"""
make_excel_template.py
===============================================================================
Build FUTAS_Excel_Template.xlsx — the spreadsheet companion to the FUTAS engine.

Deliverable #5 of the dissertation brief: "Excel formulas — High; Low; Current
Price; K; P = Low + (High - Low) x K; Signal; SL; TP; Risk/Reward."

The workbook is a *live calculator*: you type High, Low, Current Price and the
directional bias (taken from the FUTAS app verdict), and pure spreadsheet
formulas reproduce — cell for cell — what futas_engine.generate_signal() does:

    * the 15 Fibonacci Urvin levels  P = Low + (High - Low) * K
    * Stop-Loss  = the next FU level *beyond* the anchored support/resistance
    * TP1/2/3    = the next FU levels in the trade direction
    * Risk/Reward and the WAIT validity-gate (downgrade if RR < minimum)

Every SL / TP value is therefore guaranteed to be a member of the 15 FU levels,
exactly as the scientific constraint requires. Trend / market-structure
detection itself stays in the Python engine; the sheet consumes its verdict.

Run:  python make_excel_template.py
Needs: openpyxl, pandas, numpy, and futas_engine.py in the same folder.
This file is a research/algorithm-testing instrument. It is NOT financial advice.
"""
from __future__ import annotations

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import futas_engine as fe

# ----------------------------------------------------------------------------- palette
NAVY = "1f2d3d"
BLUE = "2962ff"
LBLUE = "e8eefc"
GREY = "f3f6fb"
GREEN = "1a9850"
RED = "d73027"
LINE = "c7d0db"

F_TITLE = Font(name="Calibri", size=15, bold=True, color="ffffff")
F_SUB = Font(name="Calibri", size=10, italic=True, color="ffffff")
F_HEAD = Font(name="Calibri", size=10, bold=True, color="ffffff")
F_SECT = Font(name="Calibri", size=11, bold=True, color=NAVY)
F_LBL = Font(name="Calibri", size=10, bold=True, color="1a1a1a")
F_VAL = Font(name="Calibri", size=10, color="1a1a1a")
F_IN = Font(name="Calibri", size=11, bold=True, color=BLUE)
F_NOTE = Font(name="Calibri", size=9, italic=True, color="555555")

FILL_TITLE = PatternFill("solid", fgColor=NAVY)
FILL_BLUE = PatternFill("solid", fgColor=BLUE)
FILL_HEAD = PatternFill("solid", fgColor=NAVY)
FILL_IN = PatternFill("solid", fgColor=LBLUE)
FILL_GREY = PatternFill("solid", fgColor=GREY)

THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

PRICE_FMT = "#,##0.00"
K_FMT = "0.0000"
RR_FMT = "0.00"
PCT_FMT = '0.00"%"'

# the 15 fixed coefficients straight from the engine (single source of truth)
COEFFS = list(fe.FU_COEFFICIENTS)
N = len(COEFFS)  # 15

# levels-table block on the Calculator sheet
LVL_TOP = 14                 # first coefficient row
LVL_BOT = LVL_TOP + N - 1    # == 28
LVL_RANGE = f"$C${LVL_TOP}:$C${LVL_BOT}"


# ----------------------------------------------------------------------------- helpers
def _zone(k: float) -> str:
    if k > 1.0:
        return "extension_up"
    if k < 0.0:
        return "extension_down"
    return "inside"


def _cell(ws, coord, value=None, *, font=None, fill=None, align=None,
          border=False, fmt=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = BORDER
    if fmt:
        c.number_format = fmt
    return c


def _authoritative_example():
    """Run the real engine on the shipped XAUUSD sample so the static
    'Worked_Example' sheet and the Calculator defaults are provably correct."""
    here = os.path.dirname(os.path.abspath(__file__))
    csv = os.path.join(here, "sample_data", "XAUUSD_daily.csv")
    df = fe.normalize_ohlc(pd.read_csv(csv))
    res = fe.analyze(df, asset="XAUUSD")
    return res


# ----------------------------------------------------------------------------- Calculator sheet
def build_calculator(wb: Workbook, res) -> None:
    ws = wb.active
    ws.title = "Calculator"
    ws.sheet_view.showGridLines = False
    widths = {"A": 34, "B": 16, "C": 22, "D": 13, "E": 15, "F": 14, "G": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---- title band
    ws.merge_cells("A1:F1")
    _cell(ws, "A1", "FUTAS — Fibonacci Urvin Adaptive Trading Calculator",
          font=F_TITLE, fill=FILL_TITLE, align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:F2")
    _cell(ws, "A2",
          "Scientific formula:  P = Low + (High − Low) × K        |        "
          "15 fixed Fibonacci Urvin coefficients        |        research tool, not financial advice",
          font=F_SUB, fill=FILL_TITLE, align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 16

    # ---- INPUTS
    _cell(ws, "A4", "INPUTS  (type these — High/Low come from the data, Direction from the app)",
          font=F_SECT)
    rows = [
        ("A5", "High", "B5", round(res.high, 2), PRICE_FMT),
        ("A6", "Low", "B6", round(res.low, 2), PRICE_FMT),
        ("A7", "Current Price (entry)", "B7", round(res.current_price, 2), PRICE_FMT),
    ]
    for albl, lbl, bcell, val, fmt in rows:
        _cell(ws, albl, lbl, font=F_LBL, align=LEFT, border=True, fill=FILL_GREY)
        _cell(ws, bcell, val, font=F_IN, align=RIGHT, border=True, fill=FILL_IN, fmt=fmt)

    _cell(ws, "A8", "Range = High − Low", font=F_LBL, align=LEFT, border=True, fill=FILL_GREY)
    _cell(ws, "B8", "=B5-B6", font=F_VAL, align=RIGHT, border=True, fmt=PRICE_FMT)

    _cell(ws, "A9", "Direction  (BUY / SELL / WAIT)", font=F_LBL, align=LEFT, border=True, fill=FILL_GREY)
    _cell(ws, "B9", res.signal.action, font=F_IN, align=CENTER, border=True, fill=FILL_IN)
    dv = DataValidation(type="list", formula1='"BUY,SELL,WAIT"', allow_blank=False)
    dv.error = "Choose BUY, SELL or WAIT (the verdict from the FUTAS app)."
    dv.prompt = "Pick the directional bias produced by the FUTAS engine."
    ws.add_data_validation(dv)
    dv.add(ws["B9"])

    _cell(ws, "A10", "Minimum Risk/Reward (TP1 gate)", font=F_LBL, align=LEFT, border=True, fill=FILL_GREY)
    _cell(ws, "B10", 1.0, font=F_IN, align=RIGHT, border=True, fill=FILL_IN, fmt=RR_FMT)

    _cell(ws, "A11", "Tolerance % of range (info only)", font=F_LBL, align=LEFT, border=True, fill=FILL_GREY)
    _cell(ws, "B11", 0.06, font=F_VAL, align=RIGHT, border=True, fmt="0.000")

    # ---- LEVELS TABLE
    _cell(ws, "A13", "THE 15 FIBONACCI URVIN LEVELS", font=F_SECT)
    heads = ["#", "K", "P = Low+(High−Low)×K", "% of range", "Zone", "Role vs price"]
    for j, h in enumerate(heads):
        col = get_column_letter(1 + j)
        _cell(ws, f"{col}{LVL_TOP - 1}", h, font=F_HEAD, fill=FILL_HEAD,
              align=CENTER, border=True)

    for i, k in enumerate(COEFFS):
        r = LVL_TOP + i
        _cell(ws, f"A{r}", i + 1, font=F_VAL, align=CENTER, border=True)
        _cell(ws, f"B{r}", round(k, 4), font=F_VAL, align=CENTER, border=True, fmt=K_FMT)
        # the core scientific formula, live:
        _cell(ws, f"C{r}", f"=$B$6+($B$5-$B$6)*B{r}", font=F_VAL, align=RIGHT,
              border=True, fmt=PRICE_FMT)
        _cell(ws, f"D{r}", f"=B{r}*100", font=F_VAL, align=RIGHT, border=True, fmt=PCT_FMT)
        _cell(ws, f"E{r}", f'=IF(B{r}>1,"extension_up",IF(B{r}<0,"extension_down","inside"))',
              font=F_VAL, align=CENTER, border=True)
        _cell(ws, f"F{r}",
              f'=IF(ABS(C{r}-$B$7)<0.000001,"at-price",IF(C{r}<$B$7,"support","resistance"))',
              font=F_VAL, align=CENTER, border=True)

    # ---- SIGNAL & RISK
    s = LVL_BOT + 2          # section header row (== 30)
    _cell(ws, f"A{s}", "SIGNAL & RISK — auto-selected ONLY from the 15 FU levels above",
          font=F_SECT)

    def pair(r, label, formula, fmt=PRICE_FMT, big=False):
        _cell(ws, f"A{r}", label, font=F_LBL, align=LEFT, border=True, fill=FILL_GREY)
        _cell(ws, f"B{r}", formula, font=(F_IN if big else F_VAL),
              align=RIGHT, border=True, fmt=fmt)

    e = s + 1                # first data row of the signal block
    # cell addresses (kept as names for the formulas below)
    ENTRY, ANCHOR, SL = f"B{e}", f"B{e+1}", f"B{e+2}"
    TP1, TP2, TP3 = f"B{e+3}", f"B{e+4}", f"B{e+5}"
    RISK = f"B{e+6}"
    RR1, RR2, RR3 = f"B{e+7}", f"B{e+8}", f"B{e+9}"

    pair(e, "Entry (current price)", '=IF($B$9="WAIT","—",$B$7)')
    pair(e + 1, "Anchor (support if BUY / resistance if SELL)",
         f'=IF($B$9="BUY",MAXIFS({LVL_RANGE},{LVL_RANGE},"<="&$B$7),'
         f'IF($B$9="SELL",MINIFS({LVL_RANGE},{LVL_RANGE},">="&$B$7),"—"))')
    pair(e + 2, "Stop-Loss  (next FU level beyond anchor)",
         f'=IF($B$9="BUY",IF(COUNTIFS({LVL_RANGE},"<"&{ANCHOR})>0,'
         f'MAXIFS({LVL_RANGE},{LVL_RANGE},"<"&{ANCHOR}),"n/a"),'
         f'IF($B$9="SELL",IF(COUNTIFS({LVL_RANGE},">"&{ANCHOR})>0,'
         f'MINIFS({LVL_RANGE},{LVL_RANGE},">"&{ANCHOR}),"n/a"),"—"))', big=True)
    pair(e + 3, "Take-Profit 1  (next FU level)",
         f'=IF($B$9="BUY",IF(COUNTIFS({LVL_RANGE},">"&$B$7)>0,'
         f'MINIFS({LVL_RANGE},{LVL_RANGE},">"&$B$7),"n/a"),'
         f'IF($B$9="SELL",IF(COUNTIFS({LVL_RANGE},"<"&$B$7)>0,'
         f'MAXIFS({LVL_RANGE},{LVL_RANGE},"<"&$B$7),"n/a"),"—"))')
    pair(e + 4, "Take-Profit 2",
         f'=IF($B$9="BUY",IF(AND(ISNUMBER({TP1}),COUNTIFS({LVL_RANGE},">"&{TP1})>0),'
         f'MINIFS({LVL_RANGE},{LVL_RANGE},">"&{TP1}),"n/a"),'
         f'IF($B$9="SELL",IF(AND(ISNUMBER({TP1}),COUNTIFS({LVL_RANGE},"<"&{TP1})>0),'
         f'MAXIFS({LVL_RANGE},{LVL_RANGE},"<"&{TP1}),"n/a"),"—"))')
    pair(e + 5, "Take-Profit 3",
         f'=IF($B$9="BUY",IF(AND(ISNUMBER({TP2}),COUNTIFS({LVL_RANGE},">"&{TP2})>0),'
         f'MINIFS({LVL_RANGE},{LVL_RANGE},">"&{TP2}),"n/a"),'
         f'IF($B$9="SELL",IF(AND(ISNUMBER({TP2}),COUNTIFS({LVL_RANGE},"<"&{TP2})>0),'
         f'MAXIFS({LVL_RANGE},{LVL_RANGE},"<"&{TP2}),"n/a"),"—"))')
    pair(e + 6, "Risk per unit  (|Entry − SL|)",
         f'=IF($B$9="BUY",IF(ISNUMBER({SL}),{ENTRY}-{SL},"n/a"),'
         f'IF($B$9="SELL",IF(ISNUMBER({SL}),{SL}-{ENTRY},"n/a"),"—"))')
    pair(e + 7, "Risk/Reward → TP1",
         f'=IF(AND(ISNUMBER({TP1}),ISNUMBER({RISK}),{RISK}>0),'
         f'IF($B$9="BUY",({TP1}-{ENTRY})/{RISK},IF($B$9="SELL",({ENTRY}-{TP1})/{RISK},"—")),"n/a")',
         fmt=RR_FMT)
    pair(e + 8, "Risk/Reward → TP2",
         f'=IF(AND(ISNUMBER({TP2}),ISNUMBER({RISK}),{RISK}>0),'
         f'IF($B$9="BUY",({TP2}-{ENTRY})/{RISK},IF($B$9="SELL",({ENTRY}-{TP2})/{RISK},"—")),"n/a")',
         fmt=RR_FMT)
    pair(e + 9, "Risk/Reward → TP3",
         f'=IF(AND(ISNUMBER({TP3}),ISNUMBER({RISK}),{RISK}>0),'
         f'IF($B$9="BUY",({TP3}-{ENTRY})/{RISK},IF($B$9="SELL",({ENTRY}-{TP3})/{RISK},"—")),"n/a")',
         fmt=RR_FMT)

    # FINAL SIGNAL (validity gate — mirrors generate_signal's WAIT downgrade)
    fr = e + 10
    _cell(ws, f"A{fr}", "FINAL SIGNAL  (validity-gated)", font=F_LBL, align=LEFT,
          border=True, fill=FILL_BLUE)
    ws[f"A{fr}"].font = Font(name="Calibri", size=11, bold=True, color="ffffff")
    _cell(ws, f"B{fr}",
          f'=IF($B$9="WAIT","WAIT",IF(OR(NOT(ISNUMBER({SL})),NOT(ISNUMBER({TP1})),'
          f'NOT(ISNUMBER({RR1})),{RR1}<$B$10),"WAIT (downgraded)",$B$9))',
          font=Font(name="Calibri", size=12, bold=True, color="ffffff"),
          align=CENTER, border=True, fill=FILL_BLUE)

    rr = fr + 1
    ws.merge_cells(f"A{rr}:F{rr}")
    _cell(ws, f"A{rr}",
          f'=IF($B$9="WAIT","No directional bias supplied — set Direction to the app verdict.",'
          f'IF(NOT(ISNUMBER({SL})),"Downgraded: no FU Stop-Loss beyond the anchor at this price.",'
          f'IF(NOT(ISNUMBER({TP1})),"Downgraded: no FU Take-Profit in the trade direction.",'
          f'IF({RR1}<$B$10,"Downgraded: Risk/Reward for TP1 ("&TEXT({RR1},"0.00")&") below minimum "&TEXT($B$10,"0.00")&".",'
          f'"Valid set-up: SL and TP1/2/3 are all members of the 15 FU levels."))))',
          font=F_NOTE, align=LEFT, border=True)
    ws.row_dimensions[rr].height = 28

    # footnote
    fn = rr + 2
    ws.merge_cells(f"A{fn}:F{fn}")
    _cell(ws, f"A{fn}",
          "How it works:  the anchor is the FU level the price sits on (MAXIFS ≤ price for BUY, "
          "MINIFS ≥ price for SELL); the Stop-Loss is the very next FU level beyond it; TP1/2/3 are the "
          "next FU levels in the trade direction. Identical to futas_engine.generate_signal(). "
          "Needs Excel 2019+/Microsoft 365 or LibreOffice Calc for MAXIFS/MINIFS. Open the file once "
          "so the workbook recalculates.",
          font=F_NOTE, align=LEFT)
    ws.row_dimensions[fn].height = 46

    ws.freeze_panes = "A4"


# ----------------------------------------------------------------------------- Coefficients sheet
def build_coefficients(wb: Workbook) -> None:
    ws = wb.create_sheet("Coefficients")
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 6, "B": 12, "C": 13, "D": 16, "E": 30}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    _cell(ws, "A1", "The 15 Fibonacci Urvin Adaptive Coefficients (the scientific constant)",
          font=F_TITLE, fill=FILL_TITLE, align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A2:E2")
    _cell(ws, "A2", "Order is fixed by the dissertation research. Every level: P = Low + (High − Low) × K.",
          font=F_SUB, fill=FILL_TITLE, align=Alignment(horizontal="left", vertical="center"))

    for j, h in enumerate(["#", "K", "% of range", "Zone", "Label"]):
        col = get_column_letter(1 + j)
        _cell(ws, f"{col}4", h, font=F_HEAD, fill=FILL_HEAD, align=CENTER, border=True)

    for i, k in enumerate(COEFFS):
        r = 5 + i
        _cell(ws, f"A{r}", i + 1, font=F_VAL, align=CENTER, border=True)
        _cell(ws, f"B{r}", round(k, 4), font=F_VAL, align=CENTER, border=True, fmt=K_FMT)
        _cell(ws, f"C{r}", round(k * 100.0, 2), font=F_VAL, align=CENTER, border=True, fmt=PCT_FMT)
        _cell(ws, f"D{r}", _zone(k), font=F_VAL, align=CENTER, border=True)
        _cell(ws, f"E{r}", f"FU {k:g} ({k * 100:.2f}%)", font=F_VAL, align=LEFT, border=True)

    note = 5 + N + 1
    ws.merge_cells(f"A{note}:E{note}")
    _cell(ws, f"A{note}",
          "Integrity signature (must stay identical in any published build):  "
          + ", ".join(f"{k:g}" for k in COEFFS),
          font=F_NOTE, align=LEFT)
    ws.row_dimensions[note].height = 30


# ----------------------------------------------------------------------------- Worked example sheet
def build_worked_example(wb: Workbook, res) -> None:
    ws = wb.create_sheet("Worked_Example")
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 30, "B": 18, "C": 18, "D": 18, "E": 18}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    _cell(ws, "A1", "Worked Example — static reference (computed by futas_engine.analyze)",
          font=F_TITLE, fill=FILL_TITLE, align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A2:E2")
    _cell(ws, "A2",
          "Source: sample_data/XAUUSD_daily.csv. The live 'Calculator' sheet must reproduce these "
          "numbers when fed the same High/Low/Current/Direction.",
          font=F_SUB, fill=FILL_TITLE, align=Alignment(horizontal="left", vertical="center"))

    sig = res.signal
    facts = [
        ("Asset", res.asset),
        ("Bars analysed", res.n_bars),
        ("High", round(res.high, 2)),
        ("Low", round(res.low, 2)),
        ("Range", round(res.range_size, 2)),
        ("Current Price (entry)", round(res.current_price, 2)),
        ("Trend", res.trend),
        ("Phase", res.phase),
        ("Signal", sig.action),
        ("Confidence", f"{sig.confidence} ({sig.confidence_score:.2f})"),
    ]
    r = 4
    for lbl, val in facts:
        _cell(ws, f"A{r}", lbl, font=F_LBL, align=LEFT, border=True, fill=FILL_GREY)
        _cell(ws, f"B{r}", val, font=F_VAL, align=RIGHT, border=True,
              fmt=PRICE_FMT if isinstance(val, float) else None)
        r += 1

    r += 1
    _cell(ws, f"A{r}", "Risk management — every value is a FU level", font=F_SECT)
    r += 1
    for j, h in enumerate(["", "Price", "Δ from entry", "Risk/Reward"]):
        _cell(ws, f"{get_column_letter(1 + j)}{r}", h, font=F_HEAD, fill=FILL_HEAD,
              align=CENTER, border=True)
    r += 1
    entry = res.current_price
    _cell(ws, f"A{r}", "Entry", font=F_LBL, align=LEFT, border=True)
    _cell(ws, f"B{r}", round(entry, 2), font=F_VAL, align=RIGHT, border=True, fmt=PRICE_FMT)
    _cell(ws, f"C{r}", 0.0, font=F_VAL, align=RIGHT, border=True, fmt=PRICE_FMT)
    _cell(ws, f"D{r}", "—", font=F_VAL, align=CENTER, border=True)
    r += 1
    if sig.stop_loss is not None:
        _cell(ws, f"A{r}", "Stop-Loss", font=F_LBL, align=LEFT, border=True)
        _cell(ws, f"B{r}", round(sig.stop_loss, 2), font=F_VAL, align=RIGHT, border=True, fmt=PRICE_FMT)
        _cell(ws, f"C{r}", round(sig.stop_loss - entry, 2), font=F_VAL, align=RIGHT, border=True, fmt=PRICE_FMT)
        _cell(ws, f"D{r}", "—", font=F_VAL, align=CENTER, border=True)
        r += 1
    for i, tp in enumerate(sig.take_profits):
        rr = sig.rr[i] if i < len(sig.rr) else None
        _cell(ws, f"A{r}", f"Take-Profit {i + 1}", font=F_LBL, align=LEFT, border=True)
        _cell(ws, f"B{r}", round(tp, 2), font=F_VAL, align=RIGHT, border=True, fmt=PRICE_FMT)
        _cell(ws, f"C{r}", round(tp - entry, 2), font=F_VAL, align=RIGHT, border=True, fmt=PRICE_FMT)
        _cell(ws, f"D{r}", (round(rr, 2) if rr is not None else "—"),
              font=F_VAL, align=RIGHT if rr is not None else CENTER, border=True,
              fmt=RR_FMT if rr is not None else None)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    _cell(ws, f"A{r}",
          "Verification: each SL/TP price above is one of the 15 numbers on the 'Coefficients' sheet "
          "evaluated at this High/Low — proving SL and TP are selected only from the FU levels.",
          font=F_NOTE, align=LEFT)
    ws.row_dimensions[r].height = 30


# ----------------------------------------------------------------------------- Notes sheet
def build_notes(wb: Workbook) -> None:
    ws = wb.create_sheet("How_to_use")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 110
    ws.merge_cells("A1:A1")
    _cell(ws, "A1", "How to use this workbook", font=F_TITLE, fill=FILL_TITLE,
          align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 24

    lines = [
        "",
        "1.  Open the 'Calculator' sheet. Type four inputs only:",
        "       • High  (B5)  — the high of the analysed range (the FUTAS app detects it automatically)",
        "       • Low   (B6)  — the low of the analysed range",
        "       • Current Price (B7) — the live price; this is the Entry",
        "       • Direction (B9) — BUY / SELL / WAIT, copied from the FUTAS app's verdict",
        "",
        "2.  Everything else is computed by formulas:",
        "       • Column C builds the 15 levels with  P = Low + (High − Low) × K.",
        "       • Stop-Loss is the next FU level *beyond* the support (BUY) or resistance (SELL).",
        "       • TP1/TP2/TP3 are the next FU levels in the trade direction.",
        "       • Risk/Reward = reward ÷ risk for each TP.",
        "       • FINAL SIGNAL downgrades to WAIT if no SL/TP exists or if RR(TP1) < the minimum (B10).",
        "",
        "3.  Why a spreadsheet *and* a Python engine?",
        "       The trend, swing and market-structure detection (fractals, BOS/CHoCH, regression slope)",
        "       live in futas_engine.py because they need historical bars. The spreadsheet reproduces the",
        "       deterministic part — turning the High/Low/Direction verdict into FU levels, SL, TP and RR —",
        "       so a reviewer can audit the risk arithmetic by hand, with no code.",
        "",
        "4.  Scientific guarantee:  no Entry/SL/TP number is ever typed in or copied from an old analysis.",
        "       Entry is the live price; SL and every TP are selected only from the 15 calculated FU levels.",
        "",
        "5.  Requirements:  MAXIFS / MINIFS need Excel 2019, Microsoft 365, or LibreOffice Calc 5.2+.",
        "       openpyxl writes the formulas but does not pre-compute them — open the file once to recalc.",
        "",
        "This workbook is a scientific research and algorithm-testing instrument. It does NOT provide",
        "financial advice and must not be used as the sole basis for any trading decision.",
    ]
    for i, ln in enumerate(lines):
        c = _cell(ws, f"A{i + 2}", ln, font=(F_SECT if ln[:2].strip().isdigit() and ln.strip()[1:2] == "." else F_VAL),
                  align=LEFT)
    # disclaimer emphasis
    ws[f"A{len(lines) + 1}"].font = Font(name="Calibri", size=10, bold=True, color=RED)


# ----------------------------------------------------------------------------- main
def main() -> str:
    res = _authoritative_example()
    wb = Workbook()
    build_calculator(wb, res)
    build_coefficients(wb)
    build_worked_example(wb, res)
    build_notes(wb)

    here = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(here, "FUTAS_Excel_Template.xlsx")
    wb.save(out)
    return out


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    path = main()
    print("wrote", path)
